#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PT事業部 ダッシュボード更新サーバー v2
========================================
設計方針：
  毎回アポイントリストから当月分を全て集計し直す（差分加算なし）
  前回のpt_data.jsonからは「目標・設定値・チャートラベル」のみ参照

エンドポイント：
  POST /update  : 4ファイルを受け取り集計・pt_data.jsonを返す
  GET  /health  : サーバー稼働確認用
"""

import io
import json
import os
import re
import traceback
import zipfile
from datetime import date, timedelta

import pandas as pd
import requests
from flask import Flask, jsonify, request
from flask_cors import CORS
from openpyxl import load_workbook
from supabase import create_client, Client

app = Flask(__name__)
CORS(app)

# ============================================================
# 定数
# ============================================================
SITE_LABEL  = {'新宿SC': '新宿SC', '在宅G': 'リモートSC', 'AI': 'AI'}
# 6月以降用（新宿SC→六本木SC）
SITE_LABEL_JUNE = {'新宿SC': '六本木SC', '在宅G': 'リモートSC', 'AI': 'AI'}
B_TO_D      = {'B0000106': 'D0000295', 'B0000107': 'D0000326', 'D0001318': 'B0000095'}
KONO        = '幸野有希子CRM'   # 全体/サイト集計から除外
EXCLUDE_OPS = ['堀川璃歩']      # 全集計から除外
# 在籍カウント除外スタッフ（名前が空・ランクなし・退職者等）
EXCLUDE_ENROLL = {'堀川璃歩', '堀川', ' 坂本杏奈1208', '藤公誉1212', '幸野有希子', '加藤隆治'}
VALID_RANKS    = {'トップ', 'ミドル', 'ルーキープラス', 'ルーキー', '管理者', '社員'}

# Supabaseクライアント
SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY', '')
_supabase_client = None

def get_supabase() -> Client:
    global _supabase_client
    if not _supabase_client and SUPABASE_URL and SUPABASE_KEY:
        _supabase_client = create_client(SUPABASE_URL, SUPABASE_KEY)
    return _supabase_client

# GitHubからスタッフマスターを自動読み込み
MASTER_URL = 'https://raw.githubusercontent.com/minakawa-star/digross-server/main/%E3%82%B9%E3%82%BF%E3%83%83%E3%83%95%E3%83%9E%E3%82%B9%E3%82%BF%E3%83%BC.xlsx'
_master_cache = None

def get_master_from_github():
    global _master_cache
    try:
        res = requests.get(MASTER_URL, timeout=30)
        res.raise_for_status()
        _master_cache = res.content
        return res.content
    except Exception as e:
        if _master_cache:
            return _master_cache
        raise ValueError(f'スタッフマスターの取得に失敗しました: {e}')


# ============================================================
# ヘルスチェック
# ============================================================
@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'message': 'PT事業部ダッシュボードサーバー稼働中'})


# ============================================================
# スタッフ管理エンドポイント
# ============================================================
@app.route('/staff', methods=['GET'])
def get_staff():
    try:
        sb = get_supabase()
        if not sb:
            return jsonify({'error': 'Supabase未設定'}), 500
        res = sb.table('staff').select('*').order('site').order('rank').order('name').execute()
        return jsonify({'status': 'ok', 'data': res.data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/staff', methods=['POST'])
def add_staff():
    try:
        sb = get_supabase()
        if not sb:
            return jsonify({'error': 'Supabase未設定'}), 500
        body = request.get_json()
        res = sb.table('staff').insert(body).execute()
        return jsonify({'status': 'ok', 'data': res.data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/staff/<int:staff_id>', methods=['PUT'])
def update_staff(staff_id):
    try:
        sb = get_supabase()
        if not sb:
            return jsonify({'error': 'Supabase未設定'}), 500
        body = request.get_json()
        res = sb.table('staff').update(body).eq('id', staff_id).execute()
        return jsonify({'status': 'ok', 'data': res.data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/staff/<int:staff_id>', methods=['DELETE'])
def delete_staff(staff_id):
    try:
        sb = get_supabase()
        if not sb:
            return jsonify({'error': 'Supabase未設定'}), 500
        res = sb.table('staff').delete().eq('id', staff_id).execute()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# インセンティブ管理
@app.route('/incentives/<month>', methods=['GET'])
def get_incentives(month):
    try:
        sb = get_supabase()
        if not sb:
            return jsonify({'error': 'Supabase未設定'}), 500
        res = sb.table('incentives').select('*').eq('month', month).execute()
        return jsonify({'status': 'ok', 'data': res.data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/incentives', methods=['POST'])
def upsert_incentive():
    try:
        sb = get_supabase()
        if not sb:
            return jsonify({'error': 'Supabase未設定'}), 500
        body = request.get_json()
        res = sb.table('incentives').upsert(body, on_conflict='employee_id,month').execute()
        return jsonify({'status': 'ok', 'data': res.data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================
# メイン更新エンドポイント
# ============================================================
@app.route('/update', methods=['POST'])
def update():
    try:
        # --- ファイル受け取り ---
        for key, label in [('apo','アポイントリスト'),('prod','生産性レポート'),
                            ('work','勤務データ'),('report','レポート')]:
            if key not in request.files:
                return jsonify({'error': f'{label}が見つかりません'}), 400

        apo_file  = request.files['apo'].read()
        prod_file = request.files['prod'].read()
        work_raw  = request.files['work'].read()
        work_file, work_end_date = _extract_work_from_zip(work_raw)
        prev_json = json.loads(request.files['prev'].read().decode('utf-8'))

        # スタッフマスター：アップロード優先、なければGitHubから取得
        master_file = (request.files['master'].read()
                       if 'master' in request.files
                       else get_master_from_github())

        # インセンティブ（任意）
        inc_map = prev_json.get('incentive', {})
        if 'inc' in request.files:
            inc_map = _load_incentive(request.files['inc'].read())

        # --- マスター読み込み ---
        df_master, df_wage = _load_master(master_file)
        site_map      = dict(zip(df_master['スタッフ名'], df_master['サイト']))
        rank_map      = dict(zip(df_master['スタッフ名'], df_master['ランク']))
        id_map        = dict(zip(df_master['スタッフ名'], df_master['社員番号']))
        master_ids    = set(df_master['社員番号'].tolist())
        # 社員番号ベースのマッピング（集計の主キーとして使用）
        id_site_map   = dict(zip(df_master['社員番号'], df_master['サイト']))
        id_name_map   = dict(zip(df_master['社員番号'], df_master['スタッフ名']))

        # --- アポイントリスト読み込み ---
        df_apo = _load_apo(apo_file)

        # --- 生産性レポート読み込み ---
        df_prod = _load_csv(prod_file)

        # --- 勤務データ読み込み ---
        work_data = _load_work(work_file)
        df_work   = work_data['df_monthly']   # 月累計（_calc_laborに渡す）
        df_work_daily = work_data['df_daily'] # 日次（日次人件費計算用）None if monthly

        # 汎用データの場合は最大日付を終了日として使用
        if work_data['type'] == 'daily' and df_work_daily is not None and len(df_work_daily) > 0:
            work_end_date = df_work_daily['年月日'].max()  # YYYY/MM/DD形式
        # work_end_dateは_extract_work_from_zipで取得済み（月累計の場合）

        # --- 設定値（前回pt_dataから引き継ぐもの）---
        working  = prev_json['meta']['workingDays']   # 当月営業日数
        targets  = prev_json['targets']               # 目標値
        chart_labels = prev_json['chart']['labels']   # チャートラベル

        # ============================================================
        # 当月営業日の特定
        # ============================================================
        all_get_dates = sorted(df_apo['取得日'].dropna().unique())

        month_str = prev_json['meta']['month']  # 例：2026年5月
        m = re.search(r'(\d{4})年(\d+)月', month_str)
        target_year  = int(m.group(1))
        target_month = int(m.group(2))

        biz_dates = sorted([
            d for d in all_get_dates
            if str(d).startswith(f'{target_year}/{str(target_month).zfill(2)}')
        ])

        if not biz_dates:
            return jsonify({'error': '当月のアポイントデータが見つかりません'}), 400

        elapsed = len(biz_dates)

        # 前回のelapsedDaysより少ない場合は警告（アポリストが古い可能性）
        prev_elapsed = prev_json.get('meta', {}).get('elapsedDays', 0)
        if elapsed < prev_elapsed:
            return jsonify({
                'error': f'アポリストのデータが前回より少なくなっています（前回:{prev_elapsed}営業日 → 今回:{elapsed}営業日）。'
                         f'最新のアポリスト（当月含む過去3ヶ月分）をアップしてください。'
            }), 400

        # 6月以降は新宿SC→六本木SCに表示変更
        kono_excluded = target_month < 6
        site_label = SITE_LABEL_JUNE if target_month >= 6 else SITE_LABEL

        # ============================================================
        # jinjer勤務データの期間検証
        # ファイル名の終了日とアポリストの最終営業日を比較
        # 差が2営業日以上あれば警告をmetaに記録
        # ============================================================
        work_period_warning = None
        if work_end_date:
            last_biz = biz_dates[-1]  # 例：2026/05/28
            # 終了日を yyyy/mm/dd に正規化
            work_end_norm = '/'.join(p.zfill(2) for p in work_end_date.split('/'))
            if work_end_norm < last_biz:
                work_period_warning = (
                    f'jinjer勤務データの終了日（{work_end_date}）が'
                    f'アポリストの最終営業日（{last_biz}）より前です。'
                    f'月給制スタッフの人件費が過小計上になる可能性があります。'
                    f'jinjerは当月1日〜最終更新日の期間で出力してください。'
                )

        # ============================================================
        # 人件費計算（勤務データから）
        # 【修正】days_by_id も返すよう変更
        # ============================================================
        site_labor, work_by_id, labor_by_id, days_by_id = _calc_labor(
            df_work, df_master, df_wage, master_ids, working, site_label)

        # ============================================================
        # コール・稼働人数（生産性レポートから日次集計）
        # 【修正】日付フォーマットを正規化して突合
        # ============================================================
        calls_by_date = {}
        ops_by_date   = {}

        # 生産性レポートの日付列を正規化
        # 例：「2026/5/1」→「2026/05/01」、「2026-05-01」→「2026/05/01」
        def normalize_date(s):
            s = str(s).replace('-', '/').strip()
            parts = s.split('/')
            if len(parts) == 3:
                return f"{parts[0]}/{parts[1].zfill(2)}/{parts[2].zfill(2)}"
            return s

        if '日付' in df_prod.columns:
            df_prod['日付_norm'] = df_prod['日付'].apply(normalize_date)
        else:
            df_prod['日付_norm'] = ''

        # エージェント列名を特定
        agent_col = next((c for c in ['エージェント', 'エージェント名', 'Agent']
                          if c in df_prod.columns), None)

        for d in biz_dates:
            mask = df_prod['日付_norm'] == d
            calls_by_date[d] = int(df_prod[mask]['コール数'].sum()) if mask.any() else 0
            ops_by_date[d]   = int(df_prod[mask][agent_col].nunique()) if (mask.any() and agent_col) else 0

        # ============================================================
        # 日次明細
        # ============================================================
        daily = {'all': [], 'shinjuku': [], 'remote': [], 'ai': []}
        for i, date_str in enumerate(biz_dates, 1):
            d = _calc_daily(df_apo, date_str, site_map,
                            calls_by_date, ops_by_date, i, kono_excluded, id_site_map)
            for key in ['all', 'shinjuku', 'remote', 'ai']:
                daily[key].append(d[key])

        # ============================================================
        # サイト別累計
        # ============================================================
        shinjuku_label = site_label.get('新宿SC', '新宿SC')  # 5月=新宿SC, 6月=六本木SC
        inc_site = {shinjuku_label: 0, 'リモートSC': 0, 'AI': 0}
        for name, inc_total in inc_map.items():
            if inc_total <= 0:
                continue
            site = site_label.get(site_map.get(name, ''), '')
            if site in inc_site:
                inc_site[site] += round(inc_total / working * elapsed)
        inc_all = sum(inc_site.values())

        # ============================================================
        # unit計算用にoperatorsを事前に1回だけ計算（ループ内で毎回呼ぶとタイムアウト）
        # ============================================================
        _ops_for_unit = _calc_operators(
            df_apo, biz_dates, df_master, df_prod,
            work_by_id, labor_by_id, days_by_id, id_map, site_map, rank_map,
            inc_map, elapsed, working, kono_excluded, site_label)

        sites = {}
        for k in ['all', 'shinjuku', 'remote', 'ai']:
            site_jp = {'all': None, 'shinjuku': shinjuku_label,
                       'remote': 'リモートSC', 'ai': 'AI'}[k]
            last   = daily[k][-1] if daily[k] else {}
            jinjer = {
                'all':      sum(site_labor.values()),
                'shinjuku': site_labor.get(shinjuku_label, 0),
                'remote':   site_labor.get('リモートSC', 0),
                'ai':       site_labor.get('AI', 0),
            }[k]
            labor = jinjer + (inc_all if k == 'all' else inc_site.get(site_jp, 0))

            # OP個人実績の積み上げ（全体 = 各サイトの合計が成立）
            site_ops = [op for op in _ops_for_unit
                        if k == 'all' or op['site'] == site_jp]
            sales  = sum(op['sales']  for op in site_ops)
            apo    = sum(op['apo']    for op in site_ops)
            cancel = sum(op['cancel'] for op in site_ops)
            valid  = apo - cancel
            # callsはsite_ops定義後に計算（前ループの値を使わないよう順序を保証）
            calls  = sum(op.get('calls', 0) or 0 for op in site_ops) if k != 'all' else sum(r['calls'] for r in daily['all'])
            cr     = round(cancel / apo * 100, 1) if apo > 0 else 0
            ar     = round(apo / calls * 100, 2)  if calls > 0 else 0
            cost   = round(labor / sales * 100, 1) if sales > 0 else 0

            ts = sum(o['sales'] for o in site_ops if o['sales'] > 0 and o.get('days', 0) > 0)
            td = sum(o['days']  for o in site_ops if o['sales'] > 0 and o.get('days', 0) > 0)
            unit = round(ts / td) if td > 0 else 0

            sites[k] = {
                'sales': sales, 'apo': apo, 'cancel': cancel, 'valid': valid,
                'cancelRate': cr, 'labor': labor, 'costRate': cost,
                'gross': sales - labor, 'ops': last.get('ops', 0),
                'unit': unit, 'calls': calls, 'apoRate': ar,
            }

        # ============================================================
        # チャートデータ
        # ============================================================
        chart_data = [0.0] * len(chart_labels)
        for row in daily['all']:
            d_label = row['date'].replace(f'{target_year}/', '').lstrip('0').replace('/0', '/')
            if d_label in chart_labels:
                idx = chart_labels.index(d_label)
                chart_data[idx] = round(row['sales'] / 10000, 1)

        # ============================================================
        # ヒートマップ
        # ============================================================
        heatmap = _calc_heatmap(df_apo, biz_dates, sites['all']['sales'],
                                kono_excluded, id_site_map, site_label)

        # operatorsは既にunit計算で使用した_ops_for_unitを流用（重複計算を避ける）
        operators = _ops_for_unit

        # ============================================================
        # 日次×個人成果（日次明細モーダル用）
        # ============================================================
        kono_f = (df_apo['スタッフ名'] != KONO) if kono_excluded else pd.Series([True]*len(df_apo), index=df_apo.index)

        # 【汎用データ対応】日次労働時間マップを構築
        # df_work_daily が存在する場合（汎用データ）: 日次実労働時間ベース
        # 存在しない場合（月累計）: 月次÷経過日数の均等割り
        daily_labor_per_day_default = round(sites['all']['labor'] / elapsed) if elapsed > 0 else 0

        # 時給マスター（日次人件費計算用）
        wage_by_id_d  = dict(zip(df_wage['社員番号'], df_wage['時給'].astype(float)))
        note_by_id_d  = dict(zip(df_wage['社員番号'], df_wage['備考']))
        name_to_id    = dict(zip(df_master['スタッフ名'], df_master['社員番号']))

        def _daily_labor_for_staff(name, date_str):
            """スタッフの当日人件費を計算"""
            if df_work_daily is None:
                # 月累計均等割り
                op = next((o for o in operators if o['name'] == name), None)
                return round((op.get('labor',0) or 0) / elapsed) if op and elapsed > 0 else 0
            # 汎用データから当日の実労働時間を取得
            emp_id = name_to_id.get(name, '')
            lookup = B_TO_D.get(emp_id, emp_id)
            mask = (df_work_daily['従業員ID'] == lookup) | (df_work_daily['従業員ID'] == emp_id)
            rows_d = df_work_daily[mask & (df_work_daily['年月日'] == date_str)]
            h = float(rows_d['実労働h'].sum()) if len(rows_d) > 0 else 0.0
            if h <= 0:
                return 0
            wage = wage_by_id_d.get(lookup) or wage_by_id_d.get(emp_id, 0)
            note = str(note_by_id_d.get(lookup) or note_by_id_d.get(emp_id, ''))
            if not wage:
                return 0
            if '月給' in note:
                return round(float(wage) * 1.15 / working)
            return round(float(wage) * h)

        # 日次全体人件費（daily行に付与）
        def _daily_total_labor(date_str):
            if df_work_daily is None:
                return daily_labor_per_day_default
            mask = df_work_daily['年月日'] == date_str
            rows_d = df_work_daily[mask]
            total = 0
            for _, wr in rows_d.iterrows():
                emp_id = str(wr['従業員ID'])
                lookup = B_TO_D.get(emp_id, emp_id)
                h = float(wr['実労働h'])
                if h <= 0: continue
                wage = wage_by_id_d.get(lookup) or wage_by_id_d.get(emp_id, 0)
                note = str(note_by_id_d.get(lookup) or note_by_id_d.get(emp_id, ''))
                if not wage: continue
                if '月給' in note:
                    total += round(float(wage) * 1.15 / working)
                else:
                    total += round(float(wage) * h)
            return total if total > 0 else daily_labor_per_day_default

        def _daily_total_labor(date_str, site_raw_filter=None):
            """当日の人件費合計。site_raw_filterでサイト絞り込み可能。"""
            if df_work_daily is None:
                # 均等割りフォールバック
                if site_raw_filter is None:
                    return daily_labor_per_day_default
                site_labor_val = site_labor.get(site_label.get(site_raw_filter, ''), 0)
                return round(site_labor_val / elapsed) if elapsed > 0 else 0
            mask = df_work_daily['年月日'] == date_str
            rows_d = df_work_daily[mask]
            total = 0
            for _, wr in rows_d.iterrows():
                emp_id = str(wr['従業員ID'])
                lookup = B_TO_D.get(emp_id, emp_id)
                h = float(wr['実労働h'])
                if h <= 0: continue
                # サイトフィルタ
                if site_raw_filter is not None:
                    staff_site = id_site_map.get(lookup) or id_site_map.get(emp_id, '')
                    if staff_site != site_raw_filter:
                        continue
                wage = wage_by_id_d.get(lookup) or wage_by_id_d.get(emp_id, 0)
                note = str(note_by_id_d.get(lookup) or note_by_id_d.get(emp_id, ''))
                if not wage: continue
                if '月給' in note:
                    total += round(float(wage) * 1.15 / working)
                else:
                    total += round(float(wage) * h)
            if total <= 0:
                if site_raw_filter is None:
                    return daily_labor_per_day_default
                site_labor_val = site_labor.get(site_label.get(site_raw_filter, ''), 0)
                return round(site_labor_val / elapsed) if elapsed > 0 else 0
            return total

        # daily各サイトの各行にdaily_laborを付与
        site_raw_map = {'all': None, 'shinjuku': '新宿SC', 'remote': '在宅G', 'ai': 'AI'}
        for site_key_d, raw in site_raw_map.items():
            for row in daily[site_key_d]:
                row['daily_labor'] = _daily_total_labor(row['date'], raw)

        daily_ops_by_date = {}
        for date_str in biz_dates:
            df_d_get = df_apo[(df_apo['取得日'] == date_str) & kono_f].copy()
            df_d_cxl = df_apo[(df_apo['cancel_date_str'] == date_str) & kono_f].copy()

            g_apo = df_d_get.groupby('スタッフ名').agg(
                apo=('アポイントID','count'), sg=('sales','sum')).reset_index()
            g_cxl = df_d_cxl.groupby('スタッフ名').agg(
                cxl=('アポイントID','count'), sc=('sales','sum')).reset_index()
            g = g_apo.merge(g_cxl, on='スタッフ名', how='outer').fillna(0)
            g['valid'] = (g['apo'] - g['cxl']).astype(int)
            g['net']   = (g['sg']  - g['sc']).astype(int)

            proj_by_staff = df_d_get.groupby(['スタッフ名','登録案件名']).agg(
                p_apo=('アポイントID','count'), p_sales=('sales','sum')).reset_index()
            proj_cxl_by_staff = df_d_cxl.groupby(['スタッフ名','登録案件名']).agg(
                p_cxl=('アポイントID','count'), p_sc=('sales','sum')).reset_index()
            proj_m = proj_by_staff.merge(proj_cxl_by_staff, on=['スタッフ名','登録案件名'], how='left').fillna(0)
            proj_m['p_valid'] = (proj_m['p_apo'] - proj_m['p_cxl']).astype(int)

            staff_list = []
            for _, row in g.iterrows():
                name    = str(row['スタッフ名'])
                net     = int(row['net'])
                labor_d = _daily_labor_for_staff(name, date_str)
                cost_r  = round(labor_d / net * 100, 1) if net > 0 and labor_d > 0 else None
                projs   = proj_m[proj_m['スタッフ名'] == name].copy()
                proj_list = [
                    {'name': str(r['登録案件名']), 'valid': int(r['p_valid']),
                     'sales': int(r['p_sales'] - r['p_sc'])}
                    for _, r in projs.iterrows() if int(r['p_valid']) > 0
                ]
                proj_list.sort(key=lambda x: -x['valid'])
                staff_list.append({
                    'name': name, 'sales': net, 'valid': int(row['valid']),
                    'costRate': cost_r, 'projects': proj_list
                })

            staff_list.sort(key=lambda x: (-x['sales'], -x['valid']))
            daily_ops_by_date[date_str] = staff_list

        # daily['all']の各行にdaily_opsを付与
        for row in daily['all']:
            row['daily_ops'] = daily_ops_by_date.get(row['date'], [])

        # ============================================================
        # 【修正】enrollCount・activeCount を計算
        # ============================================================
        # ============================================================
        # enrollCount・activeCount を計算
        # 在籍 = マスターに存在する有効スタッフ（除外対象・ランクなし除く）
        # アクティブ = 在籍スタッフ かつ 当月生産性レポートに名前がある
        # ============================================================
        master_names = set(df_master['スタッフ名'].tolist())
        ai_excluded = not kono_excluded

        # サイト別在籍・アクティブ集計
        site_labels_list = [shinjuku_label, 'リモートSC', 'AI']
        enroll_by_site = {s: 0 for s in site_labels_list}
        active_by_site = {s: 0 for s in site_labels_list}

        enroll_ops = []
        for op in operators:
            if op['name'].strip() in EXCLUDE_ENROLL: continue
            if op['rank'] not in VALID_RANKS: continue
            if op['name'] not in master_names: continue
            if ai_excluded and op.get('site') == 'AI': continue
            enroll_ops.append(op)
            s = op.get('site', '')
            if s in enroll_by_site:
                enroll_by_site[s] += 1

        enroll_count = len(enroll_ops)

        if agent_col:
            prod_names = set(df_prod[agent_col].astype(str).str.strip().tolist())
        else:
            prod_names = set()

        active_count = 0
        for op in enroll_ops:
            is_active = op['name'] in prod_names
            if is_active:
                active_count += 1
                s = op.get('site', '')
                if s in active_by_site:
                    active_by_site[s] += 1

        # サイト別chart dataを生成
        chart_by_site = {}
        for k in ['all', 'shinjuku', 'remote', 'ai']:
            site_jp = {'all': None, 'shinjuku': shinjuku_label, 'remote': 'リモートSC', 'ai': 'AI'}[k]
            c_data = [0.0] * len(chart_labels)
            for row in daily[k]:
                d_label = row['date'].replace(f'{target_year}/', '').replace('/0', '/').lstrip('0')
                if d_label in chart_labels:
                    idx = chart_labels.index(d_label)
                    c_data[idx] = round(row['sales'] / 10000, 1)
            chart_by_site[k] = c_data

        # ============================================================
        # PT_DATA組み立て
        # ============================================================
        today = date.today().strftime('%Y/%m/%d')
        last_date = biz_dates[-1].replace(f'{target_year}/', '')
        month_label = f'{target_year}年{target_month}月'

        PT_DATA = {
            'meta': {
                'month':           month_label,
                'lastUpdate':      today,
                'lastUpdateLabel': f"{date.today().strftime('%m/%d')}（{last_date}分反映済）",
                'elapsedDays':     elapsed,
                'workingDays':     working,
                'alertText':       f"{last_date}のデータを反映済みです（累計{elapsed}営業日）。最終更新: {today}",
                'enrollCount':     enroll_count,
                'activeCount':     active_count,
                'enrollBySite':    enroll_by_site,   # サイト別在籍数
                'activeBySite':    active_by_site,   # サイト別アクティブ数
                'workPeriodWarning': work_period_warning,
            },
            'targets':   targets,
            'sites':     sites,
            'daily':     daily,
            'chart':     {
                'labels': chart_labels,
                'data':   chart_data,          # 全体（後方互換）
                'bySite': chart_by_site,       # サイト別
            },
            'heatmap':   heatmap,
            'operators': operators,
            'incentive': inc_map,
            'prev_calls': {op['name']: op['calls'] for op in operators},
        }

        # 検証
        j = json.dumps(PT_DATA, ensure_ascii=False)
        assert 'NaN'      not in j, 'NaN混入'
        assert 'Infinity' not in j, 'Infinity混入'

        return jsonify({'status': 'ok', 'data': PT_DATA})

    except Exception as e:
        return jsonify({'error': str(e), 'detail': traceback.format_exc()}), 500


# ============================================================
# ヘルパー関数
# ============================================================
def _extract_work_from_zip(file_bytes):
    """
    zip圧縮の勤務データを展開してcsvバイトを返す。
    【追加】ファイル名から期間（終了日）を抽出して返す。
    戻り値: (csv_bytes, end_date_str or None)
    """
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            csv_files = [n for n in zf.namelist() if n.lower().endswith('.csv')]
            if not csv_files:
                raise ValueError('zip内にcsvファイルが見つかりません')
            target = max(csv_files, key=lambda n: zf.getinfo(n).file_size)
            # ファイル名から終了日を抽出
            # 例：勤務データ_打刻グループ_2026_05_01_2026_05_28.csv
            import re
            try:
                fname = target.encode('cp437').decode('cp932')
            except:
                fname = target
            end_date = None
            m = re.search(r'\d{4}_\d{2}_\d{2}_(\d{4}_\d{2}_\d{2})', fname)
            if m:
                end_date = m.group(1).replace('_', '/')
            return zf.read(target), end_date
    except zipfile.BadZipFile:
        return file_bytes, None


def _load_master(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws_m = wb['スタッフマスター']
    rows_m = list(ws_m.iter_rows(values_only=True))[1:]
    df_m = pd.DataFrame(rows_m, columns=['社員番号', 'スタッフ名', 'サイト', 'ランク'])
    df_m = df_m.dropna(subset=['社員番号']).fillna('')
    df_m['社員番号'] = df_m['社員番号'].astype(str).str.strip()

    ws_w = wb['時給マスター']
    rows_w = list(ws_w.iter_rows(values_only=True))[1:]
    # 列数に応じて柔軟に対応（4列または5列）
    header_w = list(wb['時給マスター'].iter_rows(values_only=True))[0]
    col_count = sum(1 for c in header_w if c is not None)
    if col_count >= 4:
        # 先頭4列のみ使用（末尾の空列は無視）
        rows_w_trimmed = [r[:4] for r in rows_w]
        df_w = pd.DataFrame(rows_w_trimmed, columns=['社員番号', 'スタッフ名', '時給', '備考'])
    else:
        df_w = pd.DataFrame(rows_w, columns=['社員番号', 'スタッフ名', '時給', '備考'])
    df_w = df_w.dropna(subset=['社員番号']).fillna('')
    df_w['社員番号'] = df_w['社員番号'].astype(str).str.strip()
    df_w['時給'] = pd.to_numeric(df_w['時給'], errors='coerce').fillna(0).astype(int)
    return df_m, df_w


def _load_apo(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    df = df[~df['スタッフ名'].isin(EXCLUDE_OPS)].copy()
    df['スタッフ名'] = df['スタッフ名'].replace('君塚綾子', '君塚綾子1104')
    # 名称表記ゆれ正規化（スタッフ名ベースの突合精度向上のため）
    df['スタッフ名'] = df['スタッフ名'].replace('幸野有希子', '幸野有希子CRM')
    # 社員番号列の正規化（D/B/A + 7桁数字）
    df['社員番号'] = df['社員番号'].astype(str).str.strip()
    df['cancel_date_str'] = df['キャンセル受付日'].astype(str).str.strip()
    df['sales'] = pd.to_numeric(df['案件金額'], errors='coerce').fillna(0)
    return df


def _load_csv(file_bytes):
    for enc in ['utf-8-sig', 'cp932']:
        try:
            return pd.read_csv(io.BytesIO(file_bytes), encoding=enc)
        except Exception:
            continue
    raise ValueError('CSVの読み込みに失敗しました')


def _parse_hhmm(t):
    """HH:MM形式の時間文字列をfloat（時間）に変換"""
    if pd.isna(t): return 0.0
    t = str(t).strip()
    if ':' not in t: return 0.0
    parts = t.split(':')
    try:
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 else 0
        return h + m / 60.0
    except:
        return 0.0


def _load_work(file_bytes):
    """
    jinjer勤務データの読み込み。汎用データ（日次）・月累計の両形式に対応。

    【汎用データの判定条件】
    列に「*年月日」「実労働時間」が存在する場合 → 汎用データ（日次×スタッフ）
    それ以外 → 従来の月累計データ

    【戻り値】
    {
      'type': 'daily' | 'monthly',
      'df_daily': DataFrame（日次×スタッフ：従業員ID, 年月日, 実労働h）または None,
      'df_monthly': DataFrame（スタッフ別月累計：従業員ID, 総労働時間, 出勤日数）
    }
    """
    for enc in ['cp932', 'utf-8-sig', 'utf-8']:
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), encoding=enc)
            break
        except Exception:
            continue
    else:
        raise ValueError('勤務データの読み込みに失敗しました')

    # 汎用データ判定
    is_hanyo = ('*年月日' in df.columns and '実労働時間' in df.columns and '*従業員ID' in df.columns)

    if is_hanyo:
        # --- 汎用データ（日次）---
        df = df.dropna(subset=['*従業員ID'])
        df['*従業員ID'] = df['*従業員ID'].astype(str).str.strip()
        df['年月日'] = pd.to_datetime(df['*年月日'], errors='coerce')
        df['年月日str'] = df['年月日'].dt.strftime('%Y/%m/%d')
        df['実労働h'] = df['実労働時間'].apply(_parse_hhmm)

        # 日次データ（従業員ID × 日付 × 実労働h）
        df_daily = df[['*従業員ID', '年月日str', '実労働h']].copy()
        df_daily.columns = ['従業員ID', '年月日', '実労働h']
        df_daily = df_daily[df_daily['実労働h'] > 0].reset_index(drop=True)

        # 月累計集約（_calc_laborへの互換）
        monthly = df.groupby('*従業員ID').agg(
            総労働時間=('実労働h', 'sum'),
            出勤日数=('実労働h', lambda x: (x > 0).sum())
        ).reset_index().rename(columns={'*従業員ID': '従業員ID'})

        return {'type': 'daily', 'df_daily': df_daily, 'df_monthly': monthly}

    else:
        # --- 従来の月累計データ ---
        df = df.dropna(subset=['従業員ID'])
        df['総労働時間'] = pd.to_numeric(df['総労働時間'], errors='coerce').fillna(0)
        df['出勤日数']   = pd.to_numeric(df['出勤日数'],   errors='coerce').fillna(0)
        monthly = df.sort_values('総労働時間', ascending=False).drop_duplicates(subset='従業員ID')
        return {'type': 'monthly', 'df_daily': None, 'df_monthly': monthly}


def _load_incentive(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    result = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        if len(row) >= 3 and row[2]:
            name   = str(row[1]).strip() if row[1] else ''
            amount = int(pd.to_numeric(row[2], errors='coerce') or 0)
            if name and amount > 0:
                result[name] = amount
    return result


def _calc_labor(df_work, df_master, df_wage, master_ids, working, site_label=None):
    """人件費計算"""
    if site_label is None:
        site_label = SITE_LABEL
    site_map_id = dict(zip(df_master['社員番号'], df_master['サイト']))
    wage_by_id  = dict(zip(df_wage['社員番号'], df_wage['時給']))
    note_by_id  = dict(zip(df_wage['社員番号'], df_wage['備考']))

    # 【修正】月給制は出勤日数>0のみで稼働判定、時給制は従来通り
    # まず全行をループして月給/時給で条件分岐
    shinjuku_key = site_label.get('新宿SC', '新宿SC')
    site_labor  = {shinjuku_key: 0, 'リモートSC': 0, 'AI': 0}
    work_by_id  = {}
    labor_by_id = {}
    days_by_id  = {}

    for _, row in df_work.iterrows():
        emp_id = str(row['従業員ID']).strip()
        hours  = float(row['総労働時間'])
        days   = float(row['出勤日数'])
        lookup = B_TO_D.get(emp_id, emp_id)
        if lookup not in master_ids and emp_id not in master_ids:
            continue
        site_r = site_map_id.get(lookup) or site_map_id.get(emp_id, '')
        site_d = site_label.get(site_r, 'その他')
        wage   = wage_by_id.get(lookup) or wage_by_id.get(emp_id)
        note   = str(note_by_id.get(lookup) or note_by_id.get(emp_id, ''))
        if not wage:
            continue
        wage = float(wage)
        is_monthly = '月給' in note

        # 【修正】稼働判定：月給制は出勤日数>0、時給制は労働時間>0
        if is_monthly:
            if days <= 0:
                continue
            cost = round(wage * 1.15 / working * days)
        else:
            if hours <= 0:
                continue
            cost = round(wage * hours)

        site_labor[site_d] = site_labor.get(site_d, 0) + cost
        work_by_id[lookup]  = work_by_id[emp_id]  = hours
        labor_by_id[lookup] = labor_by_id[emp_id] = cost
        days_by_id[lookup]  = days_by_id[emp_id]  = days  # 【追加】出勤日数を保存

    return site_labor, work_by_id, labor_by_id, days_by_id  # 【修正】days_by_idを追加


def _calc_daily(df_apo, date_str, site_map, calls_by_date, ops_by_date, day_num,
                kono_excluded=True, id_site_map=None):
    """1営業日分の集計。社員番号ベースでサイトを判定。"""
    kono_filter = (df_apo['スタッフ名'] != KONO) if kono_excluded else pd.Series([True]*len(df_apo), index=df_apo.index)
    df_g = df_apo[
        (df_apo['取得日'] == date_str) &
        kono_filter
    ].copy()

    df_c = df_apo[
        (df_apo['cancel_date_str'] == date_str) &
        kono_filter
    ].copy()

    # サイト判定：社員番号ベース（id_site_mapあり）、なければスタッフ名ベース
    if id_site_map:
        df_g['site_raw'] = df_g['社員番号'].map(id_site_map).fillna(
            df_g['スタッフ名'].map(site_map))
        df_c['site_raw'] = df_c['社員番号'].map(id_site_map).fillna(
            df_c['スタッフ名'].map(site_map))
    else:
        df_g['site_raw'] = df_g['スタッフ名'].map(site_map)
        df_c['site_raw'] = df_c['スタッフ名'].map(site_map)

    # 当日の案件別集計（全体・TOP5）
    pg = df_g.groupby('登録案件名').agg(apo=('アポイントID','count'), sg=('sales','sum')).reset_index()
    pc = df_c.groupby('登録案件名').agg(cxl=('アポイントID','count'), sc=('sales','sum')).reset_index()
    proj = pg.merge(pc, on='登録案件名', how='left').fillna(0)
    proj['valid'] = (proj['apo'] - proj['cxl']).astype(int)
    proj['net']   = (proj['sg']  - proj['sc']).astype(int)
    proj = proj[proj['valid'] > 0].sort_values('valid', ascending=False).head(5)
    top_projects = [
        {'name': str(r['登録案件名']), 'valid': int(r['valid']), 'sales': int(r['net'])}
        for _, r in proj.iterrows()
    ]
    max_valid = top_projects[0]['valid'] if top_projects else 1

    result = {}
    for key, raw in [('all', None), ('shinjuku', '新宿SC'),
                     ('remote', '在宅G'), ('ai', 'AI')]:
        # スタッフのサイト（マスターベース）でフィルタ
        g = df_g if raw is None else df_g[df_g['site_raw'] == raw]
        c = df_c if raw is None else df_c[df_c['site_raw'] == raw]
        row = {
            'day':    f'{day_num}営業日',
            'date':   date_str,
            'sales':  int(g['sales'].sum()) - int(c['sales'].sum()),
            'apo':    len(g),
            'cancel': len(c),
            'valid':  len(g) - len(c),
            'ops':    ops_by_date.get(date_str, 0),
            'calls':  calls_by_date.get(date_str, 0) if key == 'all' else 0,
        }
        # 案件内訳はallのみ付与
        if key == 'all':
            row['top_projects'] = top_projects
            row['max_valid']    = max_valid
        result[key] = row
    return result


def _calc_heatmap(df_apo, biz_dates, total_sales, kono_excluded=True,
                  site_map=None, site_label=None):
    """案件別売上TOP10。サイト別も生成する。"""
    kono_filter = (df_apo['スタッフ名'] != KONO) if kono_excluded else pd.Series([True]*len(df_apo), index=df_apo.index)
    df_g = df_apo[df_apo['取得日'].isin(biz_dates) & kono_filter].copy()
    df_c = df_apo[df_apo['cancel_date_str'].isin(biz_dates) & kono_filter].copy()

    def _top10(dg, dc, ts):
        pg = dg.groupby('登録案件名').agg(
            apo=('アポイントID','count'), sg=('sales','sum')).reset_index()
        pc = dc.groupby('登録案件名').agg(
            cxl=('アポイントID','count'), sc=('sales','sum')).reset_index()
        proj = pg.merge(pc, on='登録案件名', how='left').fillna(0)
        proj['valid'] = (proj['apo'] - proj['cxl']).astype(int)
        proj['net']   = (proj['sg']  - proj['sc']).astype(int)
        proj = proj[proj['valid'] > 0].sort_values('net', ascending=False).head(10)
        return [
            {'rank': i+1, 'name': str(r['登録案件名']),
             'valid': int(r['valid']), 'sales': int(r['net']),
             'pct': round(r['net'] / ts * 100, 1) if ts > 0 else 0}
            for i, (_, r) in enumerate(proj.iterrows())
        ]

    result = {'all': _top10(df_g, df_c, total_sales)}

    # サイト別
    if site_map and site_label:
        df_g['site_raw'] = df_g['社員番号'].map(
            {k: v for k, v in site_map.items()}).fillna(
            df_g['スタッフ名'].map({k: v for k, v in site_map.items()}))
        df_c['site_raw'] = df_c['社員番号'].map(
            {k: v for k, v in site_map.items()}).fillna(
            df_c['スタッフ名'].map({k: v for k, v in site_map.items()}))
        for key, raw in [('shinjuku','新宿SC'), ('remote','在宅G'), ('ai','AI')]:
            dg_s = df_g[df_g['site_raw'] == raw]
            dc_s = df_c[df_c['site_raw'] == raw]
            ts_s = int(dg_s['sales'].sum()) - int(dc_s['sales'].sum())
            result[key] = _top10(dg_s, dc_s, ts_s)
    else:
        for key in ['shinjuku', 'remote', 'ai']:
            result[key] = result['all']

    return result


def _calc_operators(df_apo, biz_dates, df_master, df_prod,
                    work_by_id, labor_by_id, days_by_id, id_map,
                    site_map, rank_map, inc_map, elapsed, working,
                    kono_excluded=True, site_label=None):
    """OP個人実績"""
    if site_label is None:
        site_label = SITE_LABEL
    kono_filter_get = (df_apo['スタッフ名'] != KONO) if kono_excluded else pd.Series([True]*len(df_apo), index=df_apo.index)
    df_get = df_apo[
        df_apo['取得日'].isin(biz_dates) &
        kono_filter_get
    ].copy()

    kouryo = (
        df_apo['cancel_date_str'].str.contains('考慮', na=False) &
        df_apo['cancel_date_str'].str.extract(
            r'(\d{4}/\d+/\d+)', expand=False).isin(biz_dates)
    )
    df_cxl_op   = df_apo[df_apo['cancel_date_str'].isin(biz_dates) | kouryo].copy()
    df_cxl_dash = df_apo[df_apo['cancel_date_str'].isin(biz_dates)].copy()

    op_get    = df_get.groupby('社員番号').agg(
        apo=('アポイントID', 'count'), sg=('sales', 'sum')).reset_index()
    op_cxl_op = df_cxl_op.groupby('社員番号').agg(
        cxl_op=('アポイントID', 'count'), sc_op=('sales', 'sum')).reset_index()
    op_cxl_d  = df_cxl_dash.groupby('社員番号').agg(
        cxl_d=('アポイントID', 'count'), sc_d=('sales', 'sum')).reset_index()

    # 生産性レポートのコール集計（日付正規化済み列を使用）
    if '日付_norm' in df_prod.columns and biz_dates:
        month_prefix = biz_dates[0][:7]  # 例：2026/05
        prod_month = df_prod[df_prod['日付_norm'].str.startswith(month_prefix)]
    else:
        prod_month = df_prod
    agent_col = None
    for col in ['エージェント', 'エージェント名', 'Agent']:
        if col in prod_month.columns:
            agent_col = col
            break
    if agent_col:
        calls_total = prod_month.groupby(agent_col)['コール数'].sum().to_dict()
    else:
        calls_total = {}

    # 集計対象：マスター全員＋アポリストに存在するID
    all_ids = set(list(df_master['社員番号']) +
                  list(op_get['社員番号'].astype(str).tolist()))
    operators = []

    for emp_id in all_ids:
        emp_id = str(emp_id).strip()
        lookup = B_TO_D.get(emp_id, emp_id)
        # スタッフ名はマスターから取得
        name = dict(zip(df_master['社員番号'], df_master['スタッフ名'])).get(lookup) or \
               dict(zip(df_master['社員番号'], df_master['スタッフ名'])).get(emp_id, '')
        if not name:
            continue

        g  = op_get[op_get['社員番号'].astype(str) == emp_id]
        co = op_cxl_op[op_cxl_op['社員番号'].astype(str) == emp_id]
        cd = op_cxl_d[op_cxl_d['社員番号'].astype(str) == emp_id]
        # B_TO_Dで変換される場合も考慮
        if len(g) == 0:
            alt = B_TO_D.get(emp_id)
            if alt:
                g  = op_get[op_get['社員番号'].astype(str) == alt]
                co = op_cxl_op[op_cxl_op['社員番号'].astype(str) == alt]
                cd = op_cxl_d[op_cxl_d['社員番号'].astype(str) == alt]

        apo    = int(g['apo'].iloc[0])      if len(g)  > 0 else 0
        sg     = int(g['sg'].iloc[0])       if len(g)  > 0 else 0
        cxl_op = int(co['cxl_op'].iloc[0]) if len(co) > 0 else 0
        sc_op  = int(co['sc_op'].iloc[0])  if len(co) > 0 else 0
        cxl_d  = int(cd['cxl_d'].iloc[0]) if len(cd) > 0 else 0
        sc_d   = int(cd['sc_d'].iloc[0])   if len(cd) > 0 else 0

        net_op   = sg - sc_op
        net_dash = sg - sc_d
        calls    = int(calls_total.get(name, 0))

        lookup = B_TO_D.get(emp_id, emp_id)
        hours  = work_by_id.get(lookup)  or work_by_id.get(emp_id, 0)
        l_base = labor_by_id.get(lookup) or labor_by_id.get(emp_id, 0)
        inc_t  = inc_map.get(name, 0)
        inc_day= round(inc_t / working * elapsed) if inc_t > 0 else 0
        labor  = l_base + inc_day

        days = days_by_id.get(lookup) or days_by_id.get(emp_id)
        if days is None:
            days = round(float(hours) / 8, 1) if hours and float(hours) > 0 else 0
        else:
            days = float(days)

        ar     = round(apo / calls * 100, 1)    if calls > 0 else None
        cost_r = round(labor / net_op * 100, 1) if net_op > 0 and labor > 0 else None
        # サイト・ランク：社員番号ベースで取得（フォールバックにスタッフ名ベース）
        site_raw = dict(zip(df_master['社員番号'], df_master['サイト'])).get(lookup) or \
                   dict(zip(df_master['社員番号'], df_master['サイト'])).get(emp_id) or \
                   site_map.get(name, '')
        site_d = site_label.get(site_raw, '')
        rank_d = rank_map.get(name, '')
        unit_pd= round(net_op / days) if days > 0 and net_op > 0 else 0

        # 案件別内訳：社員番号ベースでフィルタ
        g_proj = df_get[df_get['社員番号'].astype(str) == emp_id].groupby('登録案件名').agg(
            apo=('アポイントID','count'), sales=('sales','sum')).reset_index()
        c_proj = df_cxl_dash[df_cxl_dash['社員番号'].astype(str) == emp_id].groupby('登録案件名').agg(
            cxl=('アポイントID','count'), sc=('sales','sum')).reset_index()
        proj_m  = g_proj.merge(c_proj, on='登録案件名', how='left').fillna(0)
        proj_m['valid'] = (proj_m['apo'] - proj_m['cxl']).astype(int)
        proj_m['net']   = (proj_m['sales'] - proj_m['sc']).astype(int)
        proj_m = proj_m[proj_m['valid'] > 0].sort_values('valid', ascending=False).head(10)
        projects = [
            {'name': str(r['登録案件名']), 'valid': int(r['valid']), 'sales': int(r['net'])}
            for _, r in proj_m.iterrows()
        ]

        operators.append({
            'name': name, 'site': site_d, 'rank': rank_d,
            'sales': net_op, 'sales_dash': net_dash,
            'apo': apo, 'cancel': cxl_op, 'valid': apo - cxl_op,
            'calls': calls, 'apoRate': ar,
            'workH': round(float(hours), 1) if hours else None,
            'labor': labor,
            'labor_base': l_base, 'incentive_daily': inc_day,
            'days': days, 'unitPerDay': unit_pd, 'costRate': cost_r,
            'projects': projects,
        })

    operators.sort(key=lambda x: (-x['sales'] if x['sales'] > 0
                                  else (0 if x['sales'] == 0 else 1)))
    return operators


# ============================================================
# 起動
# ============================================================
from app_staff import register_staff_routes
register_staff_routes(app)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
