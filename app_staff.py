import os
import jwt
import datetime
import calendar
import bcrypt
from flask import jsonify, request
from supabase import create_client
from openpyxl import load_workbook
from functools import wraps

SUPABASE_STAFF_URL = os.environ.get("SUPABASE_STAFF_URL")
SUPABASE_STAFF_KEY = os.environ.get("SUPABASE_STAFF_KEY")
JWT_SECRET = os.environ.get("JWT_SECRET")

supabase_staff = create_client(SUPABASE_STAFF_URL, SUPABASE_STAFF_KEY)

MASTER_PATH = os.path.join(os.path.dirname(__file__), "スタッフマスター.xlsx")

B_TO_D = {
    "B0000106": "D0000295",
    "B0000107": "D0000326",
    "D0001318": "B0000095"
}

RATE_TABLE = {
    5: {22: (2.30, 2.00), 21: (2.30, 2.00), 20: (2.45, 2.10),
        19: (2.45, 2.10), 18: (2.45, 2.10), 17: (2.60, 2.20),
        16: (2.60, 2.20), 15: (2.60, 2.20), 14: (2.60, 2.20),
        13: (2.75, 2.30), 12: (2.75, 2.30), 11: (2.75, 2.30),
        10: (2.75, 2.30), 9: (2.75, 2.30), 8: (2.95, 2.57),
        7: (2.95, 2.57), 6: (2.95, 2.57), 5: (2.95, 2.57)},
    4: {22: (2.30, 2.00), 21: (2.30, 2.00), 20: (2.45, 2.10),
        19: (2.45, 2.10), 18: (2.45, 2.10), 17: (2.60, 2.20),
        16: (2.60, 2.20), 15: (2.60, 2.20), 14: (2.60, 2.20),
        13: (2.75, 2.30), 12: (2.75, 2.30), 11: (2.75, 2.30),
        10: (2.75, 2.30), 9: (2.75, 2.30), 8: (2.95, 2.57),
        7: (2.95, 2.57), 6: (2.95, 2.57), 5: (2.95, 2.57)},
    3: {22: (2.30, 2.00), 21: (2.30, 2.00), 20: (2.45, 2.10),
        19: (2.45, 2.10), 18: (2.45, 2.10), 17: (2.60, 2.20),
        16: (2.60, 2.20), 15: (2.60, 2.20), 14: (2.60, 2.20),
        13: (2.75, 2.30), 12: (2.75, 2.30), 11: (2.75, 2.30),
        10: (2.75, 2.30), 9: (2.75, 2.30), 8: (2.95, 2.57),
        7: (2.95, 2.57), 6: (2.95, 2.57), 5: (2.95, 2.57)},
}

# インセンティブ用：休み日数→売上比重テーブル
INCENTIVE_RATE_TABLE = [
    (1, 2.30),   # 休み0-1日 → 週5
    (4, 2.45),   # 休み2-4日 → 週4
    (8, 2.60),   # 休み5-8日 → 週3
    (12, 2.75),  # 休み9-12日 → 週2
]


def load_staff_master():
    wb = load_workbook(MASTER_PATH, data_only=True)
    ws1 = wb["スタッフマスター"]
    ws2 = wb["時給マスター"]

    master = {}
    for row in ws1.iter_rows(min_row=2, values_only=True):
        staff_id, name, site, rank = row[0], row[1], row[2], row[3]
        if not staff_id or not rank:
            continue
        sid = str(staff_id).strip()
        sid = B_TO_D.get(sid, sid)
        master[sid] = {
            "name": name, "site": site, "rank": rank,
            "hourly_wage": 0, "mgmt_fee": 0,
            "work_pattern": 5, "monthly_salary": None
        }

    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        sid = str(row[0]).strip()
        sid = B_TO_D.get(sid, sid)
        wage = row[2]
        note = str(row[3]) if row[3] else ""
        if sid not in master:
            continue
        if "月給" in note:
            master[sid]["monthly_salary"] = int(str(wage).replace(",", "").strip()) if wage else 0
        else:
            master[sid]["hourly_wage"] = int(str(wage).replace(",", "").strip()) if wage else 0
        master[sid]["mgmt_fee"] = 3030 if "管理料" in note else 0

    return master


def jwt_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        if not token:
            return jsonify({"error": "認証が必要です"}), 401
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
            request.staff_id = payload.get("staff_id")
            request.role = payload.get("role")
        except jwt.ExpiredSignatureError:
            return jsonify({"error": "トークンが期限切れです"}), 401
        except jwt.InvalidTokenError:
            return jsonify({"error": "無効なトークンです"}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        if not token:
            return jsonify({"error": "認証が必要です"}), 401
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
            if payload.get("role") != "admin":
                return jsonify({"error": "管理者権限が必要です"}), 403
            request.staff_id = payload.get("staff_id")
            request.role = payload.get("role")
        except jwt.InvalidTokenError:
            return jsonify({"error": "無効なトークンです"}), 401
        return f(*args, **kwargs)
    return decorated


def calc_campaign_fb(apo_rows, campaigns):
    """
    FBキャンペーンの集計。
    戻り値:
      breakdown: { staff_id: [ {name, category, amount, count?}, ... ] }
      totals:    { staff_id: 合計金額 }
    """
    breakdown = {}
    totals = {}

    for c in campaigns:
        target_types = c.get("target_types") or []
        exclude_resend = c.get("exclude_resend", True)
        target_staff_ids = c.get("target_staff_ids") or []
        start = c.get("start_date")
        end = c.get("end_date")
        calc_type = c.get("calc_type")
        amount = c.get("amount", 0)

        if calc_type == "fixed":
            for sid in target_staff_ids:
                sid_m = B_TO_D.get(sid, sid)
                breakdown.setdefault(sid_m, []).append({
                    "name": c.get("name"),
                    "category": c.get("category"),
                    "amount": amount
                })
                totals[sid_m] = totals.get(sid_m, 0) + amount
        else:  # per_unit
            counts = {}
            for row in apo_rows:
                acq = row.get("acquired_date")
                if not acq or not start or not end:
                    continue
                if not (start <= acq <= end):
                    continue
                if target_types and row.get("apo_type") not in target_types:
                    continue
                if exclude_resend and row.get("resend_status") == "再送":
                    continue
                sid = B_TO_D.get(row["staff_id"], row["staff_id"])
                if target_staff_ids and sid not in target_staff_ids:
                    continue
                counts[sid] = counts.get(sid, 0) + 1

            for sid, cnt in counts.items():
                amt = cnt * amount
                breakdown.setdefault(sid, []).append({
                    "name": c.get("name"),
                    "category": c.get("category"),
                    "amount": amt,
                    "count": cnt
                })
                totals[sid] = totals.get(sid, 0) + amt

    return breakdown, totals


def shift_month(target_month_str, delta):
    """target_month('YYYY-MM-01')をdeltaヶ月シフトした 'YYYY-MM-01' を返す"""
    y, m, _ = map(int, target_month_str.split("-"))
    total = (y * 12 + (m - 1)) + delta
    ny = total // 12
    nm = total % 12 + 1
    return f"{ny:04d}-{nm:02d}-01"


def get_business_days(year, month, holidays_set):
    """その月の平日数(土日・祝日を除く)を計算"""
    days_in_month = calendar.monthrange(year, month)[1]
    count = 0
    for day in range(1, days_in_month + 1):
        d = datetime.date(year, month, day)
        if d.weekday() < 5:  # Mon-Fri
            if d.isoformat() not in holidays_set:
                count += 1
    return count


def register_staff_routes(app):

    @app.route("/health_staff")
    def health_staff():
        return jsonify({"status": "ok", "service": "staff-dashboard"})

    @app.route("/staff/login", methods=["POST"])
    def staff_login():
        try:
            data = request.get_json()
            login_id = data.get("login_id", "").strip()
            password = data.get("password", "").strip()
            if not login_id or not password:
                return jsonify({"error": "IDとパスワードを入力してください"}), 400
            res = supabase_staff.table("staff_master")\
                .select("*").eq("login_id", login_id).execute()
            if not res.data:
                return jsonify({"error": "IDまたはパスワードが間違っています"}), 401
            staff = res.data[0]
            if not bcrypt.checkpw(password.encode(), staff["password_hash"].encode()):
                return jsonify({"error": "IDまたはパスワードが間違っています"}), 401
            token = jwt.encode({
                "staff_id": staff["staff_id"],
                "role": staff["role"],
                "exp": datetime.datetime.utcnow() + datetime.timedelta(days=30)
            }, JWT_SECRET, algorithm="HS256")
            return jsonify({
                "status": "ok",
                "token": token,
                "staff_id": staff["staff_id"],
                "role": staff["role"],
                "name": staff["staff_name"]
            })
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/register", methods=["POST"])
    @admin_required
    def staff_register():
        try:
            data = request.get_json()
            staff_id = data.get("staff_id", "").strip()
            staff_name = data.get("staff_name", "").strip()
            login_id = data.get("login_id", "").strip()
            password = data.get("password", "").strip()
            role = data.get("role", "staff")
            if not all([staff_id, staff_name, login_id, password]):
                return jsonify({"error": "必須項目が不足しています"}), 400
            password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
            supabase_staff.table("staff_master").upsert({
                "staff_id": staff_id,
                "staff_name": staff_name,
                "login_id": login_id,
                "password_hash": password_hash,
                "role": role
            }).execute()
            return jsonify({"status": "ok", "staff_id": staff_id})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/me")
    @jwt_required
    def staff_me():
        try:
            res = supabase_staff.table("staff_master")\
                .select("staff_id,staff_name,role")\
                .eq("staff_id", request.staff_id).execute()
            if not res.data:
                return jsonify({"error": "スタッフが見つかりません"}), 404
            return jsonify({"status": "ok", "data": res.data[0]})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/staff/debug_master")
    def debug_master():
        try:
            master = load_staff_master()
            targets = ["B0000002", "B0000032", "D0000295", "D0000326", "D0001221", "D0001316"]
            result = {k: v for k, v in master.items() if k in targets}
            return jsonify(result)
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/target", methods=["GET", "POST"])
    @jwt_required
    def staff_target():
        if request.method == "GET":
            try:
                staff_id = request.args.get("staff_id")
                month = request.args.get("month")
                if not staff_id or not month:
                    return jsonify({"error": "staff_id, monthが必要です"}), 400

                if request.role != "admin" and staff_id != request.staff_id:
                    return jsonify({"error": "権限がありません"}), 403

                target_month = month + "-01"
                res = supabase_staff.table("monthly_targets")\
                    .select("*").eq("staff_id", staff_id).eq("target_month", target_month).execute()

                if res.data:
                    return jsonify({"status": "ok", "data": res.data[0]})
                else:
                    return jsonify({"status": "ok", "data": {
                        "staff_id": staff_id, "target_month": target_month,
                        "planned_work_days": 0, "is_confirmed": False, "confirmed_work_days": None
                    }})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

        else:
            try:
                data = request.get_json()
                staff_id = data.get("staff_id")
                month = data.get("month")
                planned_work_days = data.get("planned_work_days")

                if not staff_id or not month:
                    return jsonify({"error": "staff_id, monthが必要です"}), 400

                if request.role != "admin" and staff_id != request.staff_id:
                    return jsonify({"error": "権限がありません"}), 403

                target_month = month + "-01"

                if request.role != "admin":
                    existing = supabase_staff.table("monthly_targets")\
                        .select("is_confirmed").eq("staff_id", staff_id).eq("target_month", target_month).execute()
                    if existing.data and existing.data[0]["is_confirmed"]:
                        return jsonify({"error": "確定済みのため編集できません"}), 403

                supabase_staff.table("monthly_targets").upsert({
                    "staff_id": staff_id,
                    "target_month": target_month,
                    "planned_work_days": int(planned_work_days)
                }, on_conflict="staff_id,target_month").execute()

                return jsonify({"status": "ok"})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/confirm_month", methods=["POST"])
    @admin_required
    def confirm_month():
        try:
            data = request.get_json()
            month = data.get("month")
            if not month:
                return jsonify({"error": "monthが必要です"}), 400

            target_month = month + "-01"
            master = load_staff_master()

            att_res = supabase_staff.table("attendance")\
                .select("*").eq("target_month", target_month).execute()

            work_days_map = {}
            for row in att_res.data:
                sid = B_TO_D.get(row["staff_id"], row["staff_id"])
                if (row.get("work_hours") or 0) > 0:
                    work_days_map[sid] = work_days_map.get(sid, 0) + 1

            for sid in master.keys():
                days = work_days_map.get(sid, 0)
                supabase_staff.table("monthly_targets").upsert({
                    "staff_id": sid,
                    "target_month": target_month,
                    "is_confirmed": True,
                    "confirmed_work_days": days
                }, on_conflict="staff_id,target_month").execute()

            return jsonify({"status": "ok", "message": f"{month}を確定しました"})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    # ============================================================
    # FBキャンペーン管理
    # ============================================================
    @app.route("/staff/fb_campaigns", methods=["GET", "POST"])
    @admin_required
    def fb_campaigns():
        if request.method == "GET":
            try:
                res = supabase_staff.table("fb_campaigns").select("*").order("start_date", desc=True).execute()
                return jsonify({"status": "ok", "data": res.data})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
        else:
            try:
                data = request.get_json()
                required = ["name", "category", "calc_type", "start_date", "end_date", "amount"]
                for f in required:
                    if data.get(f) is None or data.get(f) == "":
                        return jsonify({"error": f"{f}は必須です"}), 400

                record = {
                    "name": data["name"],
                    "category": data["category"],
                    "calc_type": data["calc_type"],
                    "start_date": data["start_date"],
                    "end_date": data["end_date"],
                    "amount": int(data["amount"]),
                    "target_types": data.get("target_types") or [],
                    "exclude_resend": data.get("exclude_resend", True),
                    "target_staff_ids": data.get("target_staff_ids") or []
                }
                res = supabase_staff.table("fb_campaigns").insert(record).execute()
                return jsonify({"status": "ok", "data": res.data[0]})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/fb_campaigns/<int:campaign_id>", methods=["PUT", "DELETE"])
    @admin_required
    def fb_campaign_detail(campaign_id):
        if request.method == "DELETE":
            try:
                supabase_staff.table("fb_campaigns").delete().eq("id", campaign_id).execute()
                return jsonify({"status": "ok"})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
        else:
            try:
                data = request.get_json()
                required = ["name", "category", "calc_type", "start_date", "end_date", "amount"]
                for f in required:
                    if data.get(f) is None or data.get(f) == "":
                        return jsonify({"error": f"{f}は必須です"}), 400

                record = {
                    "name": data["name"],
                    "category": data["category"],
                    "calc_type": data["calc_type"],
                    "start_date": data["start_date"],
                    "end_date": data["end_date"],
                    "amount": int(data["amount"]),
                    "target_types": data.get("target_types") or [],
                    "exclude_resend": data.get("exclude_resend", True),
                    "target_staff_ids": data.get("target_staff_ids") or []
                }
                supabase_staff.table("fb_campaigns").update(record).eq("id", campaign_id).execute()
                return jsonify({"status": "ok"})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    # ============================================================
    # インセンティブ設定（対象時給範囲）
    # ============================================================
    @app.route("/staff/incentive_settings", methods=["GET", "POST"])
    @admin_required
    def incentive_settings():
        if request.method == "GET":
            try:
                res = supabase_staff.table("incentive_settings").select("*").eq("id", 1).execute()
                if res.data:
                    return jsonify({"status": "ok", "data": res.data[0]})
                else:
                    return jsonify({"status": "ok", "data": {"id": 1, "min_wage": 2100, "max_wage": 2500}})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
        else:
            try:
                data = request.get_json()
                min_wage = int(data.get("min_wage"))
                max_wage = int(data.get("max_wage"))
                if min_wage > max_wage:
                    return jsonify({"error": "下限は上限以下にしてください"}), 400
                supabase_staff.table("incentive_settings").upsert({
                    "id": 1, "min_wage": min_wage, "max_wage": max_wage
                }).execute()
                return jsonify({"status": "ok"})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    # ============================================================
    # インセンティブ用給与データ（管理本部CSV）アップロード
    # ============================================================
    @app.route("/staff/upload/incentive_payroll_json", methods=["POST"])
    @admin_required
    def upload_incentive_payroll_json():
        try:
            data = request.get_json()
            month = data.get("month")  # "YYYY-MM"
            records = data.get("records", [])
            if not month or not records:
                return jsonify({"error": "month, recordsが必要です"}), 400

            target_month = month + "-01"
            rows = []
            for r in records:
                sid = str(r.get("staff_id", "")).strip()
                if not sid:
                    continue
                sid = B_TO_D.get(sid, sid)
                rows.append({
                    "staff_id": sid,
                    "target_month": target_month,
                    "base_salary": r.get("base_salary") or 0,
                    "overtime_allowance": r.get("overtime_allowance") or 0,
                    "commute_allowance": r.get("commute_allowance") or 0
                })

            if rows:
                supabase_staff.table("incentive_payroll").upsert(rows, on_conflict="staff_id,target_month").execute()
            return jsonify({"status": "ok", "count": len(rows)})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    # ============================================================
    # メイン集計
    # ============================================================
    @app.route("/staff/summary")
    def staff_summary():
        try:
            month = request.args.get("month")
            if not month:
                return jsonify({"error": "monthパラメータが必要です"}), 400

            target_month = month + "-01"
            master = load_staff_master()

            apo_res = supabase_staff.table("appointments")\
                .select("*").eq("target_month", target_month).execute()
            apo_rows = apo_res.data

            att_res = supabase_staff.table("attendance")\
                .select("*").eq("target_month", target_month).execute()
            att_rows = att_res.data

            targets_res = supabase_staff.table("monthly_targets")\
                .select("*").eq("target_month", target_month).execute()
            targets_map = {t["staff_id"]: t for t in targets_res.data}

            campaigns_res = supabase_staff.table("fb_campaigns").select("*").execute()
            campaign_breakdown, campaign_totals = calc_campaign_fb(apo_rows, campaigns_res.data)

            # ---- インセンティブ計算用の事前データ ----
            settings_res = supabase_staff.table("incentive_settings").select("*").eq("id", 1).execute()
            if settings_res.data:
                inc_min = settings_res.data[0]["min_wage"]
                inc_max = settings_res.data[0]["max_wage"]
            else:
                inc_min, inc_max = 2100, 2500

            prior_month = shift_month(target_month, -2)  # 2ヶ月前

            prior_att_res = supabase_staff.table("attendance")\
                .select("*").eq("target_month", prior_month).execute()
            prior_work_days_map = {}
            for row in prior_att_res.data:
                sid = B_TO_D.get(row["staff_id"], row["staff_id"])
                if (row.get("work_hours") or 0) > 0:
                    prior_work_days_map[sid] = prior_work_days_map.get(sid, 0) + 1

            prior_payroll_res = supabase_staff.table("incentive_payroll")\
                .select("*").eq("target_month", prior_month).execute()
            prior_payroll_map = {}
            for row in prior_payroll_res.data:
                sid = B_TO_D.get(row["staff_id"], row["staff_id"])
                prior_payroll_map[sid] = {
                    "base_salary": row.get("base_salary") or 0,
                    "overtime_allowance": row.get("overtime_allowance") or 0,
                    "commute_allowance": row.get("commute_allowance") or 0
                }

            holidays_res = supabase_staff.table("holidays").select("holiday_date").execute()
            holidays_set = {h["holiday_date"] for h in holidays_res.data}

            py, pm, _ = map(int, prior_month.split("-"))
            prior_business_days = get_business_days(py, pm, holidays_set)

            # ---- 結果初期化 ----
            results = {}
            for sid, info in master.items():
                results[sid] = {
                    "staff_id": sid,
                    "name": info["name"],
                    "site": info["site"],
                    "rank": info["rank"],
                    "apo_amount": 0,
                    "cxl_amount": 0,
                    "fb_achievement": 0,
                    "fb_amount": 0,
                    "fb_breakdown": [],
                    "sales": 0,
                    "work_days": 0,
                    "target_achieve": 0,
                    "target_maintain": 0,
                    "achieve_rate": None,
                    "is_monthly": info["monthly_salary"] is not None,
                    "hourly_wage": info["hourly_wage"],
                    "monthly_salary": info["monthly_salary"],
                    "planned_work_days": 0,
                    "is_confirmed": False,
                    "incentive_target": None,
                    "incentive_status": "対象外",
                    "incentive_detail": None
                }

            for row in apo_rows:
                sid = B_TO_D.get(row["staff_id"], row["staff_id"])
                if sid not in results:
                    continue
                cancel = str(row.get("cancel_date") or "")
                if cancel and cancel not in ["None", ""]:
                    results[sid]["cxl_amount"] += row.get("amount", 0)
                else:
                    results[sid]["apo_amount"] += row.get("amount", 0)
                results[sid]["fb_achievement"] += row.get("fb_amount", 0)

            for row in att_rows:
                sid = B_TO_D.get(row["staff_id"], row["staff_id"])
                if sid not in results:
                    continue
                if (row.get("work_hours") or 0) > 0:
                    results[sid]["work_days"] += 1

            for sid, r in results.items():
                info = master[sid]

                # FB内訳：達成評価分
                if r["fb_achievement"] > 0:
                    r["fb_breakdown"].append({
                        "name": "達成評価FB",
                        "category": "達成評価",
                        "amount": r["fb_achievement"]
                    })

                # FB内訳：キャンペーン分
                for entry in campaign_breakdown.get(sid, []):
                    r["fb_breakdown"].append(entry)

                campaign_total = campaign_totals.get(sid, 0)
                r["fb_amount"] = r["fb_achievement"] + campaign_total
                r["sales"] = r["apo_amount"] - r["cxl_amount"] + r["fb_amount"]

                tgt = targets_map.get(sid)
                is_confirmed = tgt["is_confirmed"] if tgt else False
                r["is_confirmed"] = is_confirmed
                r["planned_work_days"] = tgt["planned_work_days"] if tgt else 0

                if is_confirmed:
                    calc_days = tgt.get("confirmed_work_days")
                    if calc_days is None:
                        calc_days = r["work_days"]
                else:
                    calc_days = tgt["planned_work_days"] if (tgt and tgt["planned_work_days"] > 0) else r["work_days"]

                # ---- 通常の達成/維持目標 ----
                if info["monthly_salary"] is not None:
                    base = info["monthly_salary"] * 1.15 + 20000
                    r["target_achieve"] = int(base / 0.40)
                    r["target_maintain"] = int(base / 0.45)
                    if r["work_days"] == 0:
                        r["work_days"] = calc_days if calc_days > 0 else 22
                else:
                    wage = info["hourly_wage"]
                    mgmt = info["mgmt_fee"]
                    pattern = info["work_pattern"]
                    days = calc_days
                    rate_row = RATE_TABLE.get(pattern, {}).get(days)
                    if rate_row:
                        base = wage * 8 + 1000 + mgmt
                        r["target_achieve"] = int(base * days * rate_row[0])
                        r["target_maintain"] = int(base * days * rate_row[1])

                if r["target_achieve"] > 0:
                    r["achieve_rate"] = round(r["sales"] / r["target_achieve"] * 100, 1)

                # ---- インセンティブ目標 ----
                # ロジック: (基本給 + 残業手当 + 非課税通勤手当) × 売上比重
                if info["monthly_salary"] is not None:
                    r["incentive_status"] = "対象外（月給制）"
                else:
                    wage = info["hourly_wage"]
                    if wage < inc_min or wage > inc_max:
                        r["incentive_status"] = "対象外"
                    else:
                        prior_work_days = prior_work_days_map.get(sid, 0)
                        rest_days = prior_business_days - prior_work_days

                        rate = None
                        for max_rest, rate_val in INCENTIVE_RATE_TABLE:
                            if rest_days <= max_rest:
                                rate = rate_val
                                break

                        payroll = prior_payroll_map.get(sid)

                        if rate is None:
                            r["incentive_status"] = "対象外（出勤実績不足）"
                        elif payroll is None:
                            r["incentive_status"] = "対象外（給与データ未登録）"
                        else:
                            payroll_total = (payroll["base_salary"]
                                              + payroll["overtime_allowance"]
                                              + payroll["commute_allowance"])
                            incentive_target = int(payroll_total * rate)

                            r["incentive_target"] = incentive_target
                            r["incentive_status"] = "ok"
                            r["incentive_detail"] = {
                                "prior_month": prior_month[:7],
                                "prior_work_days": prior_work_days,
                                "prior_business_days": prior_business_days,
                                "rest_days": rest_days,
                                "rate": rate,
                                "base_salary": payroll["base_salary"],
                                "overtime_allowance": payroll["overtime_allowance"],
                                "commute_allowance": payroll["commute_allowance"],
                                "payroll_total": payroll_total
                            }

            return jsonify({"status": "ok", "data": list(results.values())})

        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/upload/appointments_json", methods=["POST"])
    def upload_appointments_json():
        try:
            data = request.get_json()
            records = data.get("records", [])
            if not records:
                return jsonify({"error": "データがありません"}), 400
            for r in records:
                r["target_month"] = r["acquired_date"][:7] + "-01" if r.get("acquired_date") else None
            supabase_staff.table("appointments").upsert(records, on_conflict="appointment_id").execute()
            return jsonify({"status": "ok", "count": len(records)})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/upload/productivity_json", methods=["POST"])
    def upload_productivity_json():
        try:
            data = request.get_json()
            records = data.get("records", [])
            if not records:
                return jsonify({"error": "データがありません"}), 400
            for r in records:
                r["target_month"] = r["call_date"][:7] + "-01" if r.get("call_date") else None
            supabase_staff.table("productivity").upsert(records, on_conflict="staff_id,call_date").execute()
            return jsonify({"status": "ok", "count": len(records)})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/upload/attendance_json", methods=["POST"])
    def upload_attendance_json():
        try:
            data = request.get_json()
            records = data.get("records", [])
            if not records:
                return jsonify({"error": "データがありません"}), 400
            for r in records:
                r["target_month"] = r["work_date"][:7] + "-01" if r.get("work_date") else None
            supabase_staff.table("attendance").upsert(records, on_conflict="staff_id,work_date").execute()
            return jsonify({"status": "ok", "count": len(records)})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
